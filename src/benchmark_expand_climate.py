#!/usr/bin/env python3
"""Extend the climate benchmark from 26 to 40 articles (14 additional).

Rationale: three of five climate-only frames swing by 0.25-0.45 in Cohen's kappa
across identical re-runs at n=26, so no codebook change to those frames can be
validated at the current sample size.

Method mirrors benchmark_sample_x0b.py exactly:
  * source  = climate_sample_centrality.csv, filtered to
              on_topic_flag AND topic_central_flag (the X0-filtered pool)
  * exclude = the 26 articles already in the benchmark
  * draw    = proportional stratified sample by outlet_clean x year,
              RANDOM_SEED = 42, same stratified_draw implementation

Output is a blank first-pass annotation sheet: metadata + title + body, with all
label columns empty. Holistic first-pass annotation, matching the protocol used
for the original 26 — the annotator reads the article and assigns frames
directly, without working through indicators.

Nothing is pre-filled and nothing is merged into the existing benchmark; the new
articles stay in a separate file until annotated.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SAMPLED_DIR = ROOT / "data" / "sampled"
BENCHMARK_DIR = ROOT / "data" / "benchmark"
OUT_DIR = BENCHMARK_DIR

RANDOM_SEED = 42
N_ADDITIONAL = 14
CENTRALITY_FILE = "climate_sample_centrality.csv"

LABEL_COLS = [
    "conflict_present",
    "human_interest_present",
    "economic_present",
    "deservingness_present",
    "deservingness_direction",
    "responsibility_present",
    "responsibility_responsible_actor",
    "scientific_present",
    "crisis_present",
    "solutions_present",
    "victim_present",
    "skepticism_present",
    "securitization_present",
    "othering_present",
    "othering_type",
    "agency_present",
    "agency_agency_type",
    "notes",
]

YESNO = {
    "conflict_present",
    "human_interest_present",
    "economic_present",
    "deservingness_present",
    "responsibility_present",
    "scientific_present",
    "crisis_present",
    "solutions_present",
    "victim_present",
    "skepticism_present",
    "securitization_present",
    "othering_present",
    "agency_present",
}

DROPDOWNS = {
    "deservingness_direction": '"deserving,undeserving,contested,null"',
    "othering_type": '"hostile,institutional,reported,contested,null"',
    "agency_agency_type": '"individual,collective,state,corporations,none"',
}


def sep(msg: str) -> None:
    print(f"\n{'=' * 70}\n  {msg}\n{'=' * 70}")


def stratified_draw(df: pd.DataFrame, target_n: int, seed: int) -> pd.DataFrame:
    """Proportional stratified sample by outlet_clean x year.

    Copied from benchmark_sample_x0b.py so the expansion draw is identical in
    method to the original.
    """
    pool_size = len(df)
    draws = []
    for _, grp in df.groupby(["outlet_clean", "year"]):
        n_draw = max(1, round(len(grp) * target_n / pool_size))
        n_draw = min(n_draw, len(grp))
        draws.append(grp.sample(n=n_draw, random_state=seed))

    result = pd.concat(draws)
    if len(result) > target_n:
        result = result.sample(n=target_n, random_state=seed)
    elif len(result) < target_n:
        remaining = df[~df.index.isin(result.index)]
        shortfall = target_n - len(result)
        if len(remaining) >= shortfall:
            result = pd.concat(
                [result, remaining.sample(n=shortfall, random_state=seed)]
            )
    return result.reset_index(drop=True)


def load_existing_source_rows() -> set[int]:
    ids = pd.read_csv(BENCHMARK_DIR / "benchmark_ids.csv")
    return set(ids[ids.corpus == "climate"]["source_row"].astype(int))


def main() -> None:
    sep(f"Climate benchmark expansion: +{N_ADDITIONAL} articles (26 -> 40)")

    cent = pd.read_csv(SAMPLED_DIR / CENTRALITY_FILE)
    central = cent[
        (cent["on_topic_flag"] == True) & (cent["topic_central_flag"] == True)
    ].reset_index(drop=True)
    central["source_row"] = range(len(central))
    print(f"  X0-filtered climate pool      : {len(central):,}")

    already = load_existing_source_rows()
    print(f"  already in benchmark          : {len(already)}")

    pool = central[~central["source_row"].isin(already)].reset_index(drop=True)
    print(f"  eligible pool after exclusion : {len(pool):,}")

    drawn = stratified_draw(pool, N_ADDITIONAL, RANDOM_SEED)
    assert len(drawn) == N_ADDITIONAL, len(drawn)
    assert not set(drawn["source_row"]) & already, "overlap with existing benchmark"
    print(f"  drawn (seed={RANDOM_SEED})              : {len(drawn)}")

    print("\n  outlet x year distribution of the draw:")
    for (outlet, year), n in drawn.groupby(["outlet_clean", "year"]).size().items():
        print(f"    {outlet:12s} {int(year)}   {n}")
    print("\n  by outlet:")
    for outlet, n in (
        drawn.groupby("outlet_clean").size().sort_values(ascending=False).items()
    ):
        print(f"    {outlet:12s} {n}")

    # ── build the blank annotation sheet ──────────────────────────────────────
    out = pd.DataFrame(
        {
            "benchmark_idx": range(100, 100 + len(drawn)),  # continues 0-99
            "corpus": "climate",
            "outlet_clean": drawn["outlet_clean"].values,
            "year": drawn["year"].astype(int).values,
            "date": drawn["date"].values,
            "title": drawn["title"].values,
            "body": drawn["body"].values,
            "source_row": drawn["source_row"].values,
        }
    )
    for c in LABEL_COLS:
        out[c] = None

    ids_path = OUT_DIR / "benchmark_ids_climate_expansion.csv"
    out[
        ["benchmark_idx", "corpus", "outlet_clean", "year", "date", "source_row"]
    ].to_csv(ids_path, index=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "climate_expansion"
    cols = [c for c in out.columns if c != "source_row"]
    ws.append(cols)
    for row in out[cols].itertuples(index=False):
        ws.append(["" if v is None else v for v in row])

    widths = {
        "benchmark_idx": 9,
        "corpus": 9,
        "outlet_clean": 12,
        "year": 6,
        "date": 12,
        "title": 46,
        "body": 120,
        "notes": 34,
    }
    fill = PatternFill("solid", fgColor="FFF2CC")
    for i, name in enumerate(cols, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(name, 17)
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E2F3")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if name in LABEL_COLS:
            for r in range(2, len(out) + 2):
                ws.cell(row=r, column=i).fill = fill
    for r in range(2, len(out) + 2):
        for name in ("title", "body", "notes"):
            ws.cell(row=r, column=cols.index(name) + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws.freeze_panes = "C2"

    for name, formula in [(c, '"yes,no"') for c in YESNO] + list(DROPDOWNS.items()):
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(
            type="list", formula1=formula, allow_blank=True, showDropDown=False
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(out) + 1}")

    xlsx_path = OUT_DIR / "benchmark_annotation_sheet_climate_expansion.xlsx"
    wb.save(xlsx_path)

    sep("Done")
    print(f"  blank sheet : {xlsx_path}")
    print(f"  id manifest : {ids_path}")
    print(
        f"  {len(out)} articles, benchmark_idx {out.benchmark_idx.min()}-"
        f"{out.benchmark_idx.max()}, all label columns empty."
    )
    print("\n  Annotate holistically (read article -> assign frames), matching the")
    print("  protocol used for the original 26. Do not work indicator-by-indicator.")


if __name__ == "__main__":
    main()
