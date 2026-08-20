#!/usr/bin/env python3
"""Three-way disagreement review sheet: human first-pass vs Prompt B, Run 1.

Scope is Prompt B's primary run only. Column 3 is deliberately left empty for
the annotator to complete after reading the LLM's indicator answers alongside
the original article text.

Because the pipeline discarded the raw model responses
(benchmark_stability_x3.py:168 parses `raw` for JSON and drops it), Prompt B's
PART 1 prose reasoning does not exist on disk. Column 2 therefore carries the
indicator-level yes/no sub-answers that were persisted — faithful to the
recorded run, but thinner than the chain-of-thought the prompt elicited.

Nothing in this sheet is ground truth. The first-pass column is preserved
verbatim; revisions belong in column 3 so that first-pass, LLM, and
after-reasoning annotations stay separable and revision rates can be reported
as their own result.

Output: data/validity/disagreement_review_promptB_run1.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
BENCHMARK_DIR = ROOT / "data" / "benchmark"
RUNS_DIR = ROOT / "data" / "annotated" / "benchmark_stability"
SAMPLED_DIR = ROOT / "data" / "sampled"
OUT_DIR = ROOT / "data" / "validity"

HUMAN_SHEET = BENCHMARK_DIR / "benchmark_annotation_sheet.csv.numbers"
VARIANT, RUN = "B", 1

CENTRALITY_FILES = {
    "climate": "climate_sample_centrality.csv",
    "migration": "migration_sample_centrality.csv",
}

SHARED_LABELS = [
    "conflict_present",
    "human_interest_present",
    "economic_present",
    "deservingness_present",
    "responsibility_present",
    "securitization_present",
    "othering_present",
    "agency_present",
]
MIGRATION_LABELS = ["humanitarian_present", "security_present", "policy_present"]
CLIMATE_LABELS = [
    "scientific_present",
    "crisis_present",
    "solutions_present",
    "victim_present",
    "skepticism_present",
]
LABELS = SHARED_LABELS + MIGRATION_LABELS + CLIMATE_LABELS

# Conditional fields shown as context on the frame they belong to, rather than
# as disagreement rows of their own (they are free text and would always differ).
CONDITIONAL_OF = {
    "deservingness_present": "deservingness_direction",
    "responsibility_present": "responsibility_responsible_actor",
    "agency_present": "agency_agency_type",
}

BINARY = {"yes", "no"}


# ── flag heuristics ───────────────────────────────────────────────────────────
# These are review prompts, not classifications. They point at cases worth
# checking against a known open issue; they do not decide that the issue applies.

REPORTED_ENDORSED_LABELS = {
    "skepticism_present",
    "securitization_present",
    "othering_present",
    "deservingness_present",
    "conflict_present",
}

# Word-boundary matched: short tokens like "eu" and "sp" would otherwise hit
# "Europe", "euro", "respond" and similar.
PARTY_PATTERN = re.compile(
    r"\b(pvv|vvd|cda|d66|groenlinks|pvda|fvd|sp|sgp|christenunie|bbb|volt|"
    r"denk|ja21|pvdd|forum voor democratie|part(y|ies)|partij(en)?)\b",
    re.IGNORECASE,
)

FOREIGN_GOV_PATTERN = re.compile(
    r"\b(eu|european commission|european union|brussels|brussel|unhcr|un|"
    r"united nations|nato|turkey|turkish|german(y)?|france|french|ital(y|ian)|"
    r"greece|greek|poland|polish|hungar(y|ian)|belgi(um|an)|uk|britain|british|"
    r"denmark|danish|sweden|swedish|morocc(o|an)|syria(n)?|libya(n)?|"
    r"foreign government)\b",
    re.IGNORECASE,
)
CENTRALITY_NOTE_CUES = [
    "not so much",
    "mostly",
    "not explicitly",
    "not a ",
    "centrality",
    "off-topic",
    "off topic",
    "could be argued",
    "debated",
]
KANS_TERMS = ["kansrijk", "kansarm", "kansrijke", "kansarme"]


def suggest_flags(
    label: str,
    human_val: str,
    llm_val: str,
    llm_actor: str | None,
    notes: str | None,
    body: str | None,
) -> str:
    flags = []
    body_l = (body or "").lower()
    notes_l = (notes or "").lower()
    actor_l = (llm_actor or "").lower()

    # LLM detects a frame the human did not: classic reported-vs-endorsed split.
    if label in REPORTED_ENDORSED_LABELS and llm_val == "yes" and human_val == "no":
        flags.append("reported-vs-endorsed")

    # Only meaningful on the frame the actor field belongs to; on other labels
    # the actor is incidental and the flag would be noise.
    if label == "responsibility_present" and actor_l:
        if PARTY_PATTERN.search(actor_l):
            flags.append("responsible_actor gap (political party)")
        elif FOREIGN_GOV_PATTERN.search(actor_l):
            flags.append("responsible_actor gap (foreign govt)")

    if label == "othering_present":
        flags.append("othering granularity")

    if label == "deservingness_present" or any(t in body_l for t in KANS_TERMS):
        flags.append("kansrijk/kansarm calibration")

    if any(c in notes_l for c in CENTRALITY_NOTE_CUES):
        flags.append("article centrality")

    return "; ".join(dict.fromkeys(flags))


# ── loading ───────────────────────────────────────────────────────────────────


def load_human() -> pd.DataFrame:
    try:
        from numbers_parser import Document
    except ImportError:
        sys.exit("pip install numbers-parser")
    rows = Document(str(HUMAN_SHEET)).sheets[0].tables[0].rows(values_only=True)
    df = pd.DataFrame(rows[1:], columns=list(rows[0]))
    for col in df.columns:
        df[col] = df[col].map(lambda v: None if v is None else str(v).strip())
    df["benchmark_idx"] = df["benchmark_idx"].astype(float).round().astype(int)
    return df


def load_llm() -> pd.DataFrame:
    df = pd.read_csv(RUNS_DIR / f"prompt{VARIANT}_run{RUN}.csv", dtype=str)
    for col in df.columns:
        df[col] = df[col].map(lambda v: None if pd.isna(v) else str(v).strip())
    df["benchmark_idx"] = df["article_idx"].astype(int)
    return df


def load_centrality() -> dict[int, str]:
    ids = pd.read_csv(BENCHMARK_DIR / "benchmark_ids.csv")
    out: dict[int, str] = {}
    for corpus, filename in CENTRALITY_FILES.items():
        cent = pd.read_csv(SAMPLED_DIR / filename, dtype=str)
        central = cent[
            (cent.on_topic_flag == "True") & (cent.topic_central_flag == "True")
        ].reset_index(drop=True)
        for _, r in ids[ids.corpus == corpus].iterrows():
            src = int(r.source_row)
            if src < len(central):
                out[int(r.benchmark_idx)] = central.iloc[src][
                    "centrality_justification"
                ]
    return out


def normalise(value) -> str | None:
    if value is None:
        return None
    v = str(value).strip().lower()
    return v if v in BINARY else None


def format_indicators(raw: str | None) -> str:
    if not raw:
        return ""
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return str(raw)
    return "\n".join(f"[{v}]  {q}" for q, v in parsed.items())


# ── build ─────────────────────────────────────────────────────────────────────


def collect_disagreements(human, llm, centrality) -> pd.DataFrame:
    rows = []
    for _, h in human.iterrows():
        bi = h["benchmark_idx"]
        l = llm[llm.benchmark_idx == bi]
        if l.empty:
            continue
        l = l.iloc[0]
        for label in LABELS:
            hv, lv = normalise(h.get(label)), normalise(l.get(label))
            if hv is None or lv is None or hv == lv:
                continue

            cond_field = CONDITIONAL_OF.get(label)
            h_cond = h.get(cond_field) if cond_field else None
            l_cond = l.get(cond_field) if cond_field else None
            llm_actor = l.get("responsibility_responsible_actor")

            body = h.get("body") or ""
            rows.append(
                {
                    "benchmark_idx": bi,
                    "corpus": h.get("corpus"),
                    "outlet": h.get("outlet_clean"),
                    "year": str(h.get("year") or "")[:4],
                    "title": h.get("title"),
                    "label": label.replace("_present", ""),
                    "[1] my first-pass annotation": hv,
                    "[1] my original note": h.get("notes") or "",
                    "[2] LLM annotation (Prompt B, run 1)": lv,
                    "[2] LLM indicator answers": format_indicators(
                        l.get(label.replace("_present", "") + "_indicators")
                    ),
                    "[2] LLM conditional field": (
                        f"{cond_field}: {l_cond}" if cond_field else ""
                    ),
                    "[1] my conditional field": (
                        f"{cond_field}: {h_cond}" if cond_field else ""
                    ),
                    "[3] my revised annotation": "",
                    "[3] my rationale after reading": "",
                    "suggested open-issue flag": suggest_flags(
                        label, hv, lv, llm_actor, h.get("notes"), body
                    ),
                    "article opening (full text on 'article_text' sheet)": body[:400]
                    + ("…" if len(body) > 400 else ""),
                    "centrality justification (pipeline)": centrality.get(bi, ""),
                }
            )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df.sort_values(["benchmark_idx", "label"]).reset_index(drop=True)
    df.insert(0, "review_id", range(1, len(df) + 1))
    return df


HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
BLANK_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")


def style_sheet(
    ws, widths: dict[str, int], wrap_cols: set[str], blank_cols: set[str], freeze: str
) -> None:
    headers = [c.value for c in ws[1]]
    for i, name in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(name, 18)
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=THIN)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            name = headers[cell.column - 1]
            cell.alignment = Alignment(wrap_text=name in wrap_cols, vertical="top")
            if name in blank_cols:
                cell.fill = BLANK_FILL
            cell.border = Border(bottom=THIN)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    human, llm, centrality = load_human(), load_llm(), load_centrality()
    dis = collect_disagreements(human, llm, centrality)

    wb = Workbook()

    # ── readme ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "readme"
    readme = [
        ("Three-way disagreement review — Prompt B, Run 1", ""),
        ("", ""),
        (
            "Scope",
            f"Every comparable cell where the first-pass human annotation "
            f"differs from Prompt B run {RUN}. "
            f"{len(dis)} disagreements across "
            f"{dis.benchmark_idx.nunique() if not dis.empty else 0} articles.",
        ),
        ("Primary run", "Run 1, fixed a priori. Runs 2 and 3 are not reviewed here."),
        ("Column 1", "My first-pass annotation, verbatim. Never edit this column."),
        (
            "Column 2",
            "Prompt B's annotation plus the indicator-level yes/no "
            "sub-answers it recorded.",
        ),
        (
            "Column 2 caveat",
            "Prompt B's PART 1 prose reasoning was NOT persisted — "
            "benchmark_stability_x3.py parses the raw response for "
            "JSON and discards the rest. The indicator answers are "
            "what survives. This is thinner than the chain-of-thought "
            "the prompt elicited and should be described as such.",
        ),
        (
            "Column 3",
            "Blank, for completion after reading the indicator answers "
            "ALONGSIDE the original article text (sheet 'article_text'), "
            "not from the LLM's framing alone.",
        ),
        (
            "Flags",
            "Heuristic pointers to known open issues, not classifications. "
            "They suggest what to check; they do not decide that the issue "
            "applies. Overwrite freely.",
        ),
        (
            "Excluded",
            "deservingness_direction, responsibility_responsible_actor "
            "and agency_agency_type are not disagreement rows (free text / "
            "non-binary). They appear as context on their parent frame.",
        ),
        (
            "Excluded articles",
            "Cross-corpus cells on bi=19, 21, 23, 72 are outside "
            "the comparable set and do not appear here.",
        ),
        (
            "Status",
            "Nothing here is ground truth. This is a benchmark / "
            "reference-standard comparison. First-pass, LLM, and "
            "after-reasoning annotations stay separate so three-way "
            "revision rates can be reported as their own result.",
        ),
    ]
    ws.append(["field", "value"])
    for k, v in readme:
        ws.append([k, v])
    style_sheet(ws, {"field": 22, "value": 110}, {"value"}, set(), "A2")

    # ── disagreements ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("disagreements")
    if dis.empty:
        ws.append(["no disagreements found"])
    else:
        ws.append(list(dis.columns))
        for row in dis.itertuples(index=False):
            ws.append(list(row))

        widths = {
            "review_id": 9,
            "benchmark_idx": 8,
            "corpus": 11,
            "outlet": 11,
            "year": 6,
            "title": 38,
            "label": 16,
            "[1] my first-pass annotation": 13,
            "[1] my original note": 26,
            "[2] LLM annotation (Prompt B, run 1)": 14,
            "[2] LLM indicator answers": 72,
            "[2] LLM conditional field": 26,
            "[1] my conditional field": 26,
            "[3] my revised annotation": 14,
            "[3] my rationale after reading": 40,
            "suggested open-issue flag": 32,
            "article opening (full text on 'article_text' sheet)": 60,
            "centrality justification (pipeline)": 46,
        }
        wrap = {
            "title",
            "[2] LLM indicator answers",
            "[1] my original note",
            "[3] my rationale after reading",
            "suggested open-issue flag",
            "article opening (full text on 'article_text' sheet)",
            "centrality justification (pipeline)",
            "[2] LLM conditional field",
            "[1] my conditional field",
        }
        blank = {"[3] my revised annotation", "[3] my rationale after reading"}
        style_sheet(ws, widths, wrap, blank, "H2")

        col = get_column_letter(
            list(dis.columns).index("[3] my revised annotation") + 1
        )
        dv = DataValidation(
            type="list",
            formula1='"yes,no,unresolved"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.prompt = "Fill in AFTER reading the article text."
        dv.promptTitle = "Revised annotation"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{len(dis) + 1}")

    # ── article text ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("article_text")
    keep = [
        "benchmark_idx",
        "corpus",
        "outlet_clean",
        "year",
        "date",
        "title",
        "body",
        "notes",
    ]
    art = human[keep].copy()
    art["year"] = art["year"].map(lambda v: str(v or "")[:4])
    art["centrality_justification"] = art["benchmark_idx"].map(centrality)
    if not dis.empty:
        counts = dis.groupby("benchmark_idx").size()
        art["n_disagreements"] = art["benchmark_idx"].map(counts).fillna(0).astype(int)
        art = art.sort_values(
            ["n_disagreements", "benchmark_idx"], ascending=[False, True]
        )
    ws.append(list(art.columns))
    for row in art.itertuples(index=False):
        ws.append(list(row))
    style_sheet(
        ws,
        {
            "benchmark_idx": 9,
            "corpus": 11,
            "outlet_clean": 11,
            "year": 6,
            "date": 12,
            "title": 42,
            "body": 130,
            "notes": 30,
            "centrality_justification": 50,
            "n_disagreements": 10,
        },
        {"title", "body", "notes", "centrality_justification"},
        set(),
        "B2",
    )

    out = OUT_DIR / f"disagreement_review_prompt{VARIANT}_run{RUN}.xlsx"
    wb.save(out)

    # ── console summary ───────────────────────────────────────────────────────
    print(f"wrote {out}")
    if dis.empty:
        return
    print(
        f"\n{len(dis)} disagreements across "
        f"{dis.benchmark_idx.nunique()} of 100 articles\n"
    )
    print("by label:")
    for label, n in dis.label.value_counts().items():
        sub = dis[dis.label == label]
        llm_yes = (sub["[2] LLM annotation (Prompt B, run 1)"] == "yes").sum()
        print(
            f"  {label:16s} {n:3d}   LLM=yes/human=no: {llm_yes:3d}   "
            f"LLM=no/human=yes: {n - llm_yes:3d}"
        )
    print("\nby suggested flag (non-exclusive):")
    flat = dis["suggested open-issue flag"].str.split("; ").explode()
    for flag, n in flat[flat != ""].value_counts().items():
        print(f"  {flag:42s} {n:3d}")
    print(f"\n  unflagged: {(dis['suggested open-issue flag'] == '').sum()}")


if __name__ == "__main__":
    main()
