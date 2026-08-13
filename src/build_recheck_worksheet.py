#!/usr/bin/env python3
"""Build codebook_recheck_worksheet.xlsx.

Purpose: the codebook changed after the 114-article benchmark was first-pass
annotated. Five labels are affected (scientific, deservingness, responsibility,
othering, agency). This worksheet lets the original coder re-apply her own
judgement to the changed definitions.

This is NOT the same operation as the revision-1 / revision-2 adjudication
passes. Those showed the coder the model's calls and asked her to reconsider in
light of them, which is why revision 2 had to be discarded for loss of
independence. Here the instrument changed and the coder re-reads her own calls
against new definitions. The model's annotations are deliberately absent from
every sheet in this file.

Builds the worksheet only. Changes no annotation data.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
BM = ROOT / "data" / "benchmark"
OUT = ROOT / "codebook_recheck_worksheet.xlsx"

# ── indicator text, old vs new ────────────────────────────────────────────────

SCIENTIFIC_OLD_1 = (
    "Does the article cite scientific reports, studies, " "measurements, or data?"
)
SCIENTIFIC_OLD_3 = (
    "Does the article explain climate phenomena in terms of "
    "empirical causes and effects?"
)
SCIENTIFIC_NEW_1 = (
    "Does the article cite peer-reviewed research, scientific "
    "institutions, or systematic measurement? Estimates or figures "
    "produced by advocacy groups, campaign organisations, or "
    "commercial actors do NOT count."
)
SCIENTIFIC_NEW_3 = "REMOVED — indicator 3 no longer exists in the codebook."

DESERV_OLD_3 = (
    "Does the article invoke moral criteria (effort, victimhood, "
    "compliance with rules) to evaluate recipients?"
)
DESERV_NEW_3 = (
    "Does the article invoke moral criteria (effort, victimhood, "
    "compliance with rules) to evaluate whether a GROUP merits support? "
    "Admiring or sympathetic description of one individual's effort, "
    "achievement, or hardship does NOT by itself count — the moral "
    "criteria must be applied to a category of people, not to a person "
    "as an individual."
)
DESERV_NEW_5 = (
    "NEW indicator 5: Does the article describe or report policy, legal, "
    "or administrative measures that differentiate between categories of "
    "migrant by merit, prospects, or legitimacy of claim (e.g. "
    "safe-country designations, kansrijk/kansarm classifications, "
    "conditional access to benefits, housing, or procedures) — regardless "
    "of whether the article endorses that differentiation?"
)

RESP_OLD_4 = (
    "Is there an implicit or explicit accountability claim against an "
    "identifiable party?"
)
RESP_NEW_4 = (
    "Is there an explicit accountability claim against a named or "
    "institutionally identifiable party (a government, company, agency, or "
    "organisation)? Diffuse causes such as 'human activity', 'society', or "
    "'the international community' do not count."
)

OTHERING_OLD_DEF = (
    "The article constructs or reinforces a boundary between an "
    'in-group ("us") and an out-group ("them"), positioning the '
    "out-group as different, inferior, or threatening. "
    "Presence rule: present if ANY indicator = yes."
)
OTHERING_NEW_DEF = (
    "Same boundary construction, but the out-group must be defined by "
    "nationality, ethnicity, religion, or migration status. "
    "SCOPE EXCLUSIONS — NOT othering: (a) economic or geopolitical "
    "contrasts between states, firms, or institutions (Dutch industry "
    "vs foreign competitors, EU vs UK, economies competing for "
    "investment); (b) neutral demographic, statistical or "
    "administrative description of a migrant group (population counts, "
    "housing quotas, capacity figures, policy categories) UNLESS the "
    "article attaches evaluative weight to the group itself. "
    "Presence rule: indicator 1 = yes AND at least one of 2-4 = yes."
)

# ── heuristics ────────────────────────────────────────────────────────────────
# Used only to order/flag rows for attention. Nothing is hidden on their basis
# except where the task explicitly scopes a sheet.

INDIVIDUAL_PROFILE = re.compile(
    r"\(\d{2}\)|"  # "Iman (23)"
    r"\bvertelt\b|\bvertelde\b|\bzijn verhaal\b|\bhaar verhaal\b|"
    r"\bportret\b|\binterview\b|\bik ben\b|\bmijn ouders\b|\bmijn vader\b|"
    r"\bmijn moeder\b|\bzegt hij\b|\bzegt zij\b|\bdroom\b",
    re.I,
)

INSTITUTIONAL_DIFF = re.compile(
    r"kansrijk\w*|kansarm\w*|veilig(e)? land\w*|veilige landen|safe countr\w*|"
    r"statushouder\w*|verblijfsvergunning|uitgeprocedeerd\w*|"
    r"transitkamp\w*|Dublin|nareiz\w*|gezinshereniging|"
    r"voorrang|taakstelling|uitkering\w*|bijstand|toeslag\w*|"
    r"inburger\w*|asielprocedure|terugkeer\w*|"
    r"recht op|aanspraak|toegang tot (de )?(zorg|onderwijs|arbeidsmarkt|woning)",
    re.I,
)

GEO_ECON_CONTRAST = re.compile(
    r"concurrentiepositie|concurrent\w*|marktaandeel|lidstat\w*|handelsoorlog|"
    r"\bEU\b|Europese Commissie|Brussel|Verenigd Koninkrijk|brexit|"
    r"buitenland\w*|internationale? (markt|concurrentie)|"
    r"bedrijfsleven|industrie|multinational\w*",
    re.I,
)

NEUTRAL_ADMIN = re.compile(
    r"\bCBS\b|cijfers|percentage|procent|gemiddeld\w*|statistiek\w*|"
    r"taakstelling|aantal\w*|opvangplek\w*|capaciteit|prognose|"
    r"rapport|onderzoek van|register|geregistreerd",
    re.I,
)

# The new indicator 4 requires "a government, company, agency, or organisation".
# Whitelist what qualifies rather than blacklisting what doesn't, so collectives of
# persons (e.g. "expats", "dutch politicians", "Oldenzaal residents") are surfaced
# rather than slipping through an incomplete blacklist.
INSTITUTIONAL_ACTOR = re.compile(
    r"\b(government\w*|regering\w*|kabinet|state|staat|"
    r"EU|European Commission|Commission|Council|parliament|parlement|"
    r"municipalit\w+|gemeente\w*|council|province|provincie|"
    r"corporation\w*|compan\w+|bedrijf|bedrijven|firm\w*|industry|industrie|"
    r"agency|agencies|authorit\w+|inspectie|ministr\w+|ministerie|department|"
    r"organisation\w*|organization\w*|organisatie\w*|institut\w+|"
    r"part(y|ies)|partij\w*|union|bond|sector|bank\w*|fund\w*|fonds\w*|"
    r"COA|IND|KNMI|AFM|UN|UNHCR|NATO|IPCC|CBS|WODC|"
    r"school\w*|universit\w+|hospital|ziekenhu\w+|"
    r"media|press|pers|krant\w*|omroep)\b",
    re.I,
)

BINARY = {"yes", "no"}


def norm(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().lower()
    return s if s in BINARY else None


def load_all() -> pd.DataFrame:
    """All 114 first-pass annotated articles: 100 original + 14 climate expansion."""
    from numbers_parser import Document

    rows = (
        Document(str(BM / "benchmark_annotation_sheet.csv.numbers"))
        .sheets[0]
        .tables[0]
        .rows(values_only=True)
    )
    main = pd.DataFrame(rows[1:], columns=list(rows[0]))
    for c in main.columns:
        main[c] = main[c].map(lambda v: None if v is None else str(v).strip())
    main["benchmark_idx"] = main.benchmark_idx.astype(float).round().astype(int)

    exp = pd.read_csv(
        BM / "benchmark_annotation_sheet_climate_expansion - climate_expansion.csv",
        dtype=str,
    )
    exp["benchmark_idx"] = exp.benchmark_idx.astype(int)

    df = pd.concat([main, exp], ignore_index=True).sort_values("benchmark_idx")
    ids = pd.read_csv(BM / "benchmark_ids.csv")
    climate = set(ids[ids.corpus == "climate"].benchmark_idx.astype(int)) | set(
        range(100, 114)
    )
    df["corpus_group"] = df.benchmark_idx.map(
        lambda b: "climate" if b in climate else "migration"
    )
    return df.reset_index(drop=True)


def text_of(r) -> str:
    return f"{r.get('title') or ''}\n\n{r.get('body') or ''}"


# ── sheet writer ──────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="D9E2F3")
BLANK_FILL = PatternFill("solid", fgColor="FFF2CC")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")

WIDTHS = {
    "benchmark_idx": 9,
    "corpus": 10,
    "outlet": 11,
    "year": 6,
    "title": 40,
    "article text": 110,
    "subset": 22,
    "flagged for review": 17,
    "why flagged": 34,
    "my original call": 13,
    "my original notes": 26,
    "my original indicator answers": 22,
    "old indicator 1": 40,
    "old indicator 3": 40,
    "new indicator 1": 46,
    "new indicator 3": 40,
    "old indicator 3 (deservingness)": 40,
    "new indicator 3 (deservingness)": 46,
    "new indicator 5 (deservingness)": 46,
    "old indicator 4": 40,
    "new indicator 4": 46,
    "old definition + presence rule": 44,
    "new definition + presence rule": 52,
    "my original responsible_actor": 24,
    "my original agency_type": 18,
    "[REVISE] revised present": 15,
    "[REVISE] othering_type": 17,
    "[REVISE] note on why it changed": 40,
}

BLANKS = {
    "[REVISE] revised present",
    "[REVISE] othering_type",
    "[REVISE] note on why it changed",
}


def write_sheet(
    wb,
    name: str,
    rows: list[dict],
    cols: list[str],
    dropdowns: dict[str, str] | None = None,
):
    ws = wb.create_sheet(name)
    if not rows:
        ws.append([f"No articles met the filter for '{name}'."])
        return ws
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    for i, cname in enumerate(cols, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = WIDTHS.get(cname, 18)
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for ri in range(2, len(rows) + 2):
        flagged = (
            str(ws.cell(row=ri, column=cols.index("flagged for review") + 1).value)
            .strip()
            .lower()
            == "yes"
            if "flagged for review" in cols
            else False
        )
        for i, cname in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=i)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cname in BLANKS:
                cell.fill = BLANK_FILL
            elif flagged and cname == "flagged for review":
                cell.fill = FLAG_FILL

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    for cname, formula in (dropdowns or {}).items():
        if cname not in cols:
            continue
        letter = get_column_letter(cols.index(cname) + 1)
        dv = DataValidation(
            type="list", formula1=formula, allow_blank=True, showDropDown=False
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")
    return ws


BASE = ["benchmark_idx", "corpus", "outlet", "year", "title"]
TAIL = ["[REVISE] revised present", "[REVISE] note on why it changed", "article text"]
YESNO = {"[REVISE] revised present": '"yes,no"'}


def base_fields(r) -> dict:
    return {
        "benchmark_idx": int(r.benchmark_idx),
        "corpus": r.corpus_group,
        "outlet": r.get("outlet_clean"),
        "year": str(r.get("year") or "")[:4],
        "title": r.get("title"),
        "article text": text_of(r),
        "my original notes": r.get("notes") or "",
        "my original indicator answers": "not recorded — holistic first-pass",
    }


def main() -> None:
    df = load_all()
    wb = Workbook()
    wb.remove(wb.active)
    counts = {}

    # ── Sheet 1: scientific ───────────────────────────────────────────────────
    rows = []
    for _, r in df[df.corpus_group == "climate"].iterrows():
        if norm(r.get("scientific_present")) != "yes":
            continue
        rows.append(
            {
                **base_fields(r),
                "my original call": "yes",
                "old indicator 1": SCIENTIFIC_OLD_1,
                "old indicator 3": SCIENTIFIC_OLD_3,
                "new indicator 1": SCIENTIFIC_NEW_1,
                "new indicator 3": SCIENTIFIC_NEW_3,
            }
        )
    cols = (
        BASE
        + [
            "my original call",
            "my original indicator answers",
            "old indicator 1",
            "old indicator 3",
            "new indicator 1",
            "new indicator 3",
            "my original notes",
        ]
        + TAIL
    )
    write_sheet(wb, "1_scientific", rows, cols, YESNO)
    counts["1_scientific"] = len(rows)

    # ── Sheet 2: deservingness ────────────────────────────────────────────────
    rows = []
    for _, r in df.iterrows():
        call = norm(r.get("deservingness_present"))
        if call is None:
            continue
        txt = text_of(r)
        if call == "yes":
            hits = set(m.group(0).lower() for m in INDIVIDUAL_PROFILE.finditer(txt))
            flag = bool(hits)
            why = (
                "individual-profile markers: " + ", ".join(sorted(hits)[:4])
                if flag
                else ""
            )
            subset = "A. my YES calls — indicator 3 tightening"
        else:
            hits = set(m.group(0).lower() for m in INSTITUTIONAL_DIFF.finditer(txt))
            flag = bool(hits)
            why = (
                "institutional-differentiation terms: " + ", ".join(sorted(hits)[:5])
                if flag
                else ""
            )
            subset = "B. my NO calls — new indicator 5"
        if not flag:
            continue
        rows.append(
            {
                **base_fields(r),
                "subset": subset,
                "my original call": call,
                "flagged for review": "yes",
                "why flagged": why,
                "old indicator 3 (deservingness)": DESERV_OLD_3,
                "new indicator 3 (deservingness)": DESERV_NEW_3,
                "new indicator 5 (deservingness)": DESERV_NEW_5,
            }
        )
    rows.sort(key=lambda r: (r["subset"], r["benchmark_idx"]))
    cols = (
        BASE[:1]
        + ["subset"]
        + BASE[1:]
        + [
            "my original call",
            "flagged for review",
            "why flagged",
            "my original indicator answers",
            "old indicator 3 (deservingness)",
            "new indicator 3 (deservingness)",
            "new indicator 5 (deservingness)",
            "my original notes",
        ]
        + TAIL
    )
    write_sheet(wb, "2_deservingness", rows, cols, YESNO)
    counts["2_deservingness"] = len(rows)

    # ── Sheet 3: responsibility ───────────────────────────────────────────────
    rows = []
    for _, r in df.iterrows():
        if norm(r.get("responsibility_present")) != "yes":
            continue
        actor = str(r.get("responsibility_responsible_actor") or "").strip()
        if not actor or actor.lower() in ("null", "none", "nan"):
            why = "responsible_actor is empty/null despite present=yes"
        elif not INSTITUTIONAL_ACTOR.search(actor):
            why = (
                f"'{actor}' is not a government, company, agency or "
                f"organisation — a collective of persons or diffuse cause"
            )
        else:
            continue
        rows.append(
            {
                **base_fields(r),
                "my original call": "yes",
                "my original responsible_actor": actor or "(empty)",
                "flagged for review": "yes",
                "why flagged": why,
                "old indicator 4": RESP_OLD_4,
                "new indicator 4": RESP_NEW_4,
            }
        )
    cols = (
        BASE
        + [
            "my original call",
            "my original responsible_actor",
            "flagged for review",
            "why flagged",
            "my original indicator answers",
            "old indicator 4",
            "new indicator 4",
            "my original notes",
        ]
        + TAIL
    )
    write_sheet(wb, "3_responsibility", rows, cols, YESNO)
    counts["3_responsibility"] = len(rows)

    # ── Sheet 4: othering ─────────────────────────────────────────────────────
    rows = []
    for _, r in df.iterrows():
        if norm(r.get("othering_present")) != "yes":
            continue
        txt = text_of(r)
        geo = set(m.group(0).lower() for m in GEO_ECON_CONTRAST.finditer(txt))
        adm = set(m.group(0).lower() for m in NEUTRAL_ADMIN.finditer(txt))
        why = []
        if len(geo) >= 2:
            why.append(
                "possible state/firm economic-geopolitical contrast: "
                + ", ".join(sorted(geo)[:4])
            )
        if len(adm) >= 3:
            why.append(
                "possible neutral demographic/administrative description: "
                + ", ".join(sorted(adm)[:4])
            )
        rows.append(
            {
                **base_fields(r),
                "my original call": "yes",
                "flagged for review": "yes" if why else "",
                "why flagged": " | ".join(why),
                "old definition + presence rule": OTHERING_OLD_DEF,
                "new definition + presence rule": OTHERING_NEW_DEF,
            }
        )
    cols = BASE + [
        "my original call",
        "flagged for review",
        "why flagged",
        "my original indicator answers",
        "old definition + presence rule",
        "new definition + presence rule",
        "my original notes",
        "[REVISE] revised present",
        "[REVISE] othering_type",
        "[REVISE] note on why it changed",
        "article text",
    ]
    write_sheet(
        wb,
        "4_othering",
        rows,
        cols,
        {
            **YESNO,
            "[REVISE] othering_type": '"hostile,institutional,reported,contested"',
        },
    )
    counts["4_othering"] = len(rows)

    # ── Sheet 5: agency (computed) ────────────────────────────────────────────
    rows = []
    for _, r in df.iterrows():
        recorded = norm(r.get("agency_present"))
        atype = str(r.get("agency_agency_type") or "").strip().lower()
        if recorded is None or atype in ("", "nan"):
            continue
        derived = "no" if atype in ("none", "null") else "yes"
        if derived != recorded:
            rows.append(
                {
                    **base_fields(r),
                    "my original call": recorded,
                    "my original agency_type": atype,
                    "why flagged": f"recorded present='{recorded}' but "
                    f"agency_type='{atype}' derives '{derived}'",
                    "[REVISE] revised present": derived,
                }
            )
    cols = BASE + [
        "my original call",
        "my original agency_type",
        "why flagged",
        "[REVISE] revised present",
        "[REVISE] note on why it changed",
        "article text",
    ]
    ws = write_sheet(wb, "5_agency_computed", rows, cols, YESNO)
    if not rows:
        ws.cell(row=1, column=1).value = (
            "No disagreements. For all 114 articles the recorded agency_present "
            "matches the new derivation agency_present = (agency_type != 'none'). "
            "The restructuring changed the derivation mechanism, not the underlying "
            "type judgement — nothing to re-check here."
        )
        ws.column_dimensions["A"].width = 130
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    counts["5_agency_computed"] = len(rows)

    # ── readme ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("0_readme", 0)
    readme = [
        (
            "WHAT THIS IS",
            "Re-application of the ORIGINAL CODER'S OWN JUDGEMENT to updated codebook "
            "definitions. The codebook changed after all 114 articles were first-pass "
            "annotated; five labels are affected.",
        ),
        (
            "WHAT THIS IS NOT",
            "This is NOT reconsideration in light of the model's output. The LLM's "
            "annotations appear NOWHERE in this file — not its calls, not its "
            "indicators, not its counts. That is deliberate.",
        ),
        (
            "WHY THE DISTINCTION MATTERS",
            "The earlier revision-1 and revision-2 passes showed the coder the model's "
            "calls and asked her to reconsider. Revision 2 was discarded because all 11 "
            "changes moved toward the model, breaking independence. THIS pass has no "
            "such exposure: the instrument changed, and the coder re-reads her own calls "
            "against new definitions. It does not carry the same independence concern "
            "and should be documented as a distinct operation.",
        ),
        ("SCOPE", "114 articles = original 100 benchmark + 14 climate expansion."),
        ("", ""),
        ("HOW EACH SHEET WAS SCOPED", ""),
        (
            "1_scientific",
            "HARD RULE. All climate articles (n=40) where my original call was 'yes'. "
            "'no' calls are excluded by design: the indicator change only narrows what "
            "counts, so a 'no' cannot become a 'yes'. No heuristic filtering.",
        ),
        (
            "2_deservingness",
            "HEURISTIC PRE-FILTER. Two subsets. (A) my YES calls surfaced by "
            "individual-profile markers — the indicator-3 tightening targets "
            "judgements resting on one person's effort rather than on a category. "
            "(B) my NO calls surfaced by institutional-differentiation language — "
            "the new indicator 5. Tuned to over-include; borderline cases are shown.",
        ),
        (
            "3_responsibility",
            "EXACT FILTER from a recorded field. Every article where present='yes' AND "
            "the recorded responsible_actor is diffuse/non-institutional or empty. "
            "No heuristics — this is read straight off responsible_actor.",
        ),
        (
            "4_othering",
            "FULL SET, HEURISTICALLY FLAGGED. Every article where my original call was "
            "'yes' is listed. The 'flagged for review' column marks those whose text "
            "suggests a new scope exclusion applies. Unflagged rows are still shown "
            "and still need confirming under the new presence rule.",
        ),
        (
            "5_agency_computed",
            "COMPUTED, NO MANUAL INPUT NEEDED. Recomputes agency_present under the new "
            "derivation (agency_type != 'none') from recorded agency_type values and "
            "lists only disagreements with the recorded call.",
        ),
        ("", ""),
        (
            "A NOTE ON 'my original indicator answers'",
            "This column is empty on every sheet. Indicator-level answers were never "
            "recorded on the human side: the annotation protocol was holistic (read the "
            "article, assign the frame), not indicator-by-indicator. The column is kept "
            "so the absence is explicit rather than looking like missing data.",
        ),
        (
            "A NOTE ON 'my original notes'",
            "Populated for only 7 of 114 articles, so the deservingness heuristics work "
            "off article text rather than recorded rationale.",
        ),
        ("", ""),
        (
            "HOW TO FILL IT IN",
            "Yellow columns are yours. Leave '[REVISE] revised present' blank if your "
            "original call still stands under the new definition; fill it only where it "
            "changes. Use the note column to say why. On sheet 4, assign othering_type "
            "for any case you confirm as 'yes'.",
        ),
        (
            "NEXT STEP",
            "Return the completed file. Merging revised calls into a clean post-fix "
            "reference standard, running the LLM under the final codebook, and "
            "recomputing validity are all separate steps.",
        ),
    ]
    ws.append(["field", "detail"])
    for k, v in readme:
        ws.append([k, v])
    ws.append(["", ""])
    ws.append(["ROW COUNTS", "; ".join(f"{k}={v}" for k, v in counts.items())])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 118
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HDR_FILL

    wb.save(OUT)
    print(f"wrote {OUT}\n")
    for k, v in counts.items():
        print(f"  {k:22s} {v:3d} rows")


if __name__ == "__main__":
    main()
